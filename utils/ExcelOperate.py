import openpyxl
from openpyxl import load_workbook

# from CoonfigFile.VarConfig import testDataPaht


class ExcelOperate:

    def __init__(self):
        self.workbook = None
        self.sheet = None

    def load_workbook(self,filename):
        '''
        加载 excel文件
        :param filename:
        :return:
        '''
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            print(e)

    def get_sheet_by_name(self,sheetname):
        '''
        根据sheetname 获取具体sheet
        :param sheetname:
        :return:
        '''
        try:
            self.sheet = self.workbook[sheetname]
        except Exception as e:
            print(e)

    def get_rows_nums(self):
        '''
        获取行数
        :return:
        '''
        return self.sheet.max_row

    def get_col_nums(self):
        '''
        获取列数
        :return:
        '''
        return self.sheet.max_column

    def get_row_values(self,row):
        '''
        获取某一行的值
        :param row:
        :return:
        '''
        columns = self.sheet.max_column
        row_data= []
        for i in range(1,columns+1):
            cell_value = self.sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def get_cell_vlaue(self,row,column):
        '''
        获取某个单元格的值
        :param row:
        :param column:
        :return:
        '''
        cell_value = self.sheet.cell(row=row,column=column).value
        return cell_value


    def write_cell(self, row, column, value, file_name):
        '''
        写入单元格的值
        :param row:
        :param column:
        :param value:
        :return:
        '''
        try:
            self.sheet.cell(row=row, column=column, value=value)
            self.workbook.save(file_name)
        except Exception as e:
            raise e

if __name__ == '__main__':
    pass
    # ex = ExcelOperate()
    # ex.load_workbook(testDataPaht)
    # ex.get_sheet_by_name("Sheet1")
    # print(ex.get_rows_nums())
    # print(ex.get_col_nums())
    # print(ex.get_row_values(2))
    # print(ex.get_cell_vlaue(2,2))
    # ex.write_cell(2,7,'pass')